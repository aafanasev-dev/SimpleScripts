import os
import argparse
from copy import deepcopy

from openpyxl import load_workbook

def get_common_numbers(directory):
    unique_numbers = set()

    is_first_file = True

    for filename in os.listdir(directory):
        if not filename.lower().endswith(".xlsx"):
            continue

        filepath = os.path.join(directory, filename)
        workbook = load_workbook(filepath, data_only=True)
        sheet = workbook.active  # first sheet

        row = 2  # start from row 2
        file_numbers = set()
        while True:
            cell_value = sheet.cell(row=row, column=1).value  # column A

            if cell_value is None:
                break  # stop at first empty cell

            file_numbers.add(cell_value)
            row += 1

        if is_first_file:
            unique_numbers = deepcopy(file_numbers)
            is_first_file = False
        else:
            unique_numbers &= file_numbers
            if len(unique_numbers) == 0:
                break;

    return unique_numbers


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Collect common numbers from Excel files')
    parser.add_argument('-i', '--input', default=os.path.dirname(os.path.abspath(__file__)),
                        help='Input directory path (default: script directory)')
    args = parser.parse_args()

    result = get_common_numbers(args.input)

    print(f"Unique numbers: {result}")
